"""
Christian Tour Dashboard ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ data processor
Single source: CT Dashboard.xlsx

Sheet layout:
  B2C 2026 Actuals  ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ wide: branch | region | Jan..Dec  (k EUR actuals 2026)
  P                 ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ wide: branch | region | Jan..Dec  (k EUR plan 2026)
  LY                ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ wide: branch | region | Jan..Dec  (k EUR actuals 2025)
  B2C Daily         ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ long: date | branch | pax | reservations | revenue
  B2B Daily         ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ long: partner | pax | actuals | month
  B2B P             ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ monthly 2026 plan for B2B

All k-EUR values ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ 1000 on ingest ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ stored as EUR.
B2C Daily / B2B Daily values assumed to already be in EUR (not k EUR).
"""

import io
import gc
import logging
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd

logger = logging.getLogger(__name__)
_b2b_stream_diag: dict = {}  # populated by _stream_b2b_daily_direct

K_EUR = 1.0

MONTH_ABBR = {
    "ian": 1, "jan": 1, "january": 1, "ianuarie": 1,
    "feb": 2, "february": 2, "februarie": 2,
    "mar": 3, "march": 3, "martie": 3,
    "apr": 4, "april": 4, "aprilie": 4,
    "mai": 5, "may": 5,
    "iun": 6, "jun": 6, "june": 6, "iunie": 6,
    "iul": 7, "jul": 7, "july": 7, "iulie": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9, "septembrie": 9,
    "oct": 10, "october": 10, "octombrie": 10,
    "nov": 11, "november": 11, "noiembrie": 11,
    "dec": 12, "december": 12, "decembrie": 12,
}

SMALL_SHEETS  = ["B2C 2026 Actuals", "P", "LY", "B2B P"]
STREAM_SHEETS = ["B2C Daily", "B2B Daily"]
SHEETS_NEEDED = SMALL_SHEETS + STREAM_SHEETS


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ helpers ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        import math
        return float(v) if not math.isnan(float(v)) else None
    s = str(v).replace(",", ".").replace("\xa0", "").strip()
    if s in ("", "-", "ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ", "n/a", "na", "#n/a"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _col_to_month(col: Any) -> Optional[int]:
    if isinstance(col, datetime):
        return col.month
    if isinstance(col, date) and not isinstance(col, datetime):
        return col.month
    if isinstance(col, int) and 1 <= col <= 12:
        return col
    # Excel serial date (openpyxl may return raw int instead of datetime)
    if isinstance(col, (int, float)) and 45 < col < 100000:
        try:
            from datetime import timedelta
            dt = date(1899, 12, 30) + timedelta(days=int(col))
            return dt.month
        except Exception:
            pass
    s = str(col).strip().lower()[:15]
    try:
        m = int(float(s))
        if 1 <= m <= 12:
            return m
    except ValueError:
        pass
    for kw, m in MONTH_ABBR.items():
        if s.startswith(kw):
            return m
    return None


def _find_header_row(df: pd.DataFrame) -> int:
    for i, row in df.iterrows():
        str_vals = [c for c in row if isinstance(c, str) and c.strip()]
        if len(str_vals) >= 3:
            return int(i)
    return 0


def _promote_header(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    idx = _find_header_row(df)
    if idx > 0 or not any(isinstance(c, str) for c in df.columns):
        df = df.copy()
        df.columns = [str(c).strip() if c is not None else f"col_{i}"
                      for i, c in enumerate(df.iloc[idx])]
        df = df.iloc[idx + 1:].reset_index(drop=True)
    return df


def _detect_branch_col(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        s = str(col).lower()
        if any(kw in s for kw in ("branch", "agentie", "canal", "channel",
                                   "sucursala", "denumire", "name", "unitate", "punct")):
            return col
    candidates = []
    for col in df.columns:
        if df[col].dtype == object:
            fill = df[col].dropna().shape[0] / max(df.shape[0], 1)
            candidates.append((fill, col))
    return max(candidates, key=lambda x: x[0])[1] if candidates else None


def _detect_region_col(df: pd.DataFrame, branch_col: Optional[str]) -> Optional[str]:
    for col in df.columns:
        if col == branch_col:
            continue
        s = str(col).lower()
        if any(kw in s for kw in ("region", "regiune", "zona", "zone", "area")):
            return col
    return None


def _header_col_idx(header: list, *keywords, exclude: set = None) -> Optional[int]:
    """Find column index in a raw header row by keyword."""
    for i, h in enumerate(header):
        if exclude and i in exclude:
            continue
        s = str(h).lower().strip() if h is not None else ""
        if any(kw in s for kw in keywords):
            return i
    return None


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ B2C wide sheets ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def parse_b2c_wide_sheet(raw_df: pd.DataFrame, sheet_name: str,
                          field: str, year: int) -> List[Dict]:
    df = _promote_header(raw_df.copy())
    if df.empty:
        logger.warning("%s: empty after header promotion", sheet_name)
        return []

    branch_col = _detect_branch_col(df)
    region_col = _detect_region_col(df, branch_col)

    month_map: Dict[str, int] = {}
    for col in df.columns:
        m = _col_to_month(col)
        if m:
            month_map[col] = m

    if len(month_map) < 2:
        logger.warning("%s: only %d month cols ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ cols: %s",
                       sheet_name, len(month_map), list(df.columns)[:20])
        return []

    logger.info("%s: branch=%s region=%s months=%d",
                sheet_name, branch_col, region_col, len(month_map))

    records = []
    for _, row in df.iterrows():
        branch = str(row[branch_col]).strip() if branch_col and pd.notna(row.get(branch_col)) else None
        if not branch or branch.lower() in ("nan", "none", "", "-"):
            continue
        if _col_to_month(branch) is not None:
            continue
        region = str(row[region_col]).strip() if region_col and pd.notna(row.get(region_col)) else None
        for col, m in month_map.items():
            v = _num(row.get(col))
            if v is None:
                continue
            records.append({
                "sheet": sheet_name, "year": year, "month": m,
                "branch": branch, "region": region,
                field: round(v * K_EUR, 2),
            })
    logger.info("%s: %d records", sheet_name, len(records))
    return records


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ B2C Daily (openpyxl streaming) ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def _stream_b2c_daily(wb, valid_branches: Set[str]) -> List[Dict]:
    """
    Stream B2C Daily row-by-row from an open openpyxl workbook.
    Aggregates to (year, month, branch) WITHOUT building an intermediate list.
    """
    if "B2C Daily" not in wb.sheetnames:
        logger.warning("B2C Daily sheet not found")
        return []

    ws = wb["B2C Daily"]
    header = None
    date_idx = branch_idx = pax_idx = res_idx = rev_idx = None
    agg: Dict[Tuple, Dict] = {}
    norm_valid = {b.strip().lower() for b in valid_branches}
    rows_processed = 0

    for row_vals in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row_vals):
            continue

        # detect header on first non-empty row
        if header is None:
            str_count = sum(1 for v in row_vals if isinstance(v, str) and v.strip())
            if str_count >= 3:
                header = [str(v).strip().lower() if v is not None else "" for v in row_vals]
                date_idx   = _header_col_idx(header, "date", "data", "zi", "day", "datum")
                branch_idx = _header_col_idx(header, "branch", "agentie", "canal", "channel",
                                             "sucursala", "denumire", "name", "unitate", "punct")
                pax_idx    = _header_col_idx(header, "pax", "pasageri", "persons", "persoane")
                res_idx    = _header_col_idx(header, "rezervari", "reservations", "bookings")
                rev_idx    = _header_col_idx(header, "valoare", "value", "revenue",
                                             "vanzari", "sales", "incasari", "total")
                logger.info("B2C Daily header: date=%s branch=%s pax=%s res=%s rev=%s",
                            date_idx, branch_idx, pax_idx, res_idx, rev_idx)
                if branch_idx is None or date_idx is None:
                    logger.warning("B2C Daily: cannot find date/branch columns ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ skipping")
                    return []
            continue

        # parse branch
        branch = None
        if branch_idx is not None and branch_idx < len(row_vals) and row_vals[branch_idx] is not None:
            branch = str(row_vals[branch_idx]).strip()
        if not branch or branch.lower() not in norm_valid:
            continue

        # parse date
        dt = row_vals[date_idx] if date_idx < len(row_vals) else None
        if isinstance(dt, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    dt = datetime.strptime(dt.strip(), fmt)
                    break
                except ValueError:
                    dt = None
        if not hasattr(dt, "year"):
            continue

        yr, mo = dt.year, dt.month
        key = (yr, mo, branch)
        if key not in agg:
            agg[key] = {"revenue": 0.0, "pax": 0, "reservations": 0}

        if rev_idx is not None and rev_idx < len(row_vals):
            v = _num(row_vals[rev_idx])
            if v:
                agg[key]["revenue"] += v
        if pax_idx is not None and pax_idx < len(row_vals):
            v = _num(row_vals[pax_idx])
            if v:
                agg[key]["pax"] += int(v)
        if res_idx is not None and res_idx < len(row_vals):
            v = _num(row_vals[res_idx])
            if v:
                agg[key]["reservations"] += int(v)
        rows_processed += 1

    logger.info("B2C Daily: streamed %d rows ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ %d (year,month,branch) buckets",
                rows_processed, len(agg))
    return [{
        "sheet": "B2C Daily",
        "year": yr, "month": mo, "branch": branch, "region": None,
        "revenue_daily": round(vals["revenue"], 2),
        "pax": vals["pax"] or None,
        "reservations": vals["reservations"] or None,
    } for (yr, mo, branch), vals in agg.items()]


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ B2B Daily (openpyxl streaming) ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def _stream_b2b_daily(wb) -> List[Dict]:
    """Stream B2B Daily row-by-row from an open openpyxl workbook."""
    if "B2B Daily" not in wb.sheetnames:
        logger.warning("B2B Daily sheet not found")
        return []

    ws = wb["B2B Daily"]
    header = None
    partner_idx = pax_idx = rev_idx = month_idx = year_idx = None
    records = []
    skip_no_partner = skip_no_month = 0
    current_year = date.today().year

    for row_vals in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row_vals):
            continue

        if header is None:
            str_count = sum(1 for v in row_vals if isinstance(v, str) and v.strip())
            if str_count >= 2:
                candidate = [str(v).strip().lower() if v is not None else "" for v in row_vals]
                c_partner = _header_col_idx(candidate, "partner", "client", "agentie",
                                            "agency", "denumire", "name", "partener")
                c_month   = _header_col_idx(candidate, "luna", "month", "lunÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ",
                                            "period", "data", "date")
                # Only accept row as header if it has BOTH a partner and a month col
                if c_partner is not None and c_month is not None:
                    header      = candidate
                    partner_idx = c_partner
                    pax_idx     = _header_col_idx(header, "pax", "pasageri", "persons")
                    rev_idx     = _header_col_idx(header, "valoare", "value", "revenue",
                                                  "vanzari", "actual", "sales",
                                                  "incasat", "eur", "nett", "net",
                                                  "total", "realizat",
                                                  exclude={pax_idx} if pax_idx is not None else None)
                    month_idx   = c_month
                    year_idx    = _header_col_idx(header, "year", "an", "anul")
                    logger.info("B2B Daily header: partner=%s pax=%s rev=%s month=%s year=%s | cols=%s",
                                partner_idx, pax_idx, rev_idx, month_idx, year_idx,
                                [str(h)[:20] if h else None for h in row_vals[:8]])
                else:
                    logger.info("B2B Daily: skipping non-header row (no partner+month): %s",
                                [str(v)[:15] for v in row_vals[:6]])
            continue

        partner = None
        if partner_idx is not None and partner_idx < len(row_vals) and row_vals[partner_idx] is not None:
            partner = str(row_vals[partner_idx]).strip()
        if not partner or partner.lower() in ("nan", "none", "", "-"):
            skip_no_partner += 1
            continue
        # Skip numeric-only values (row numbers / blank cells stored as 0)
        if partner.replace(".", "").lstrip("-").isdigit():
            skip_no_partner += 1
            continue

        # parse month/year
        raw_m = row_vals[month_idx] if month_idx is not None and month_idx < len(row_vals) else None
        mo = yr = None

        if isinstance(raw_m, (datetime, date)):
            mo, yr = raw_m.month, raw_m.year
        elif isinstance(raw_m, (int, float)) and 45 < raw_m < 100000:
            # Excel serial date ÃÂ¢ÃÂÃÂ extract BOTH month and year
            try:
                _dt = date(1899, 12, 30) + timedelta(days=int(raw_m))
                mo, yr = _dt.month, _dt.year
            except Exception:
                pass
        elif raw_m is not None:
            mo = _col_to_month(raw_m)
            if not mo:
                for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%m/%Y", "%d/%m/%Y"):
                    try:
                        dt = datetime.strptime(str(raw_m).strip(), fmt)
                        mo, yr = dt.month, dt.year
                        break
                    except ValueError:
                        pass

        if yr is None:
            yr = current_year
            if year_idx is not None and year_idx < len(row_vals):
                yv = _num(row_vals[year_idx])
                if yv:
                    yr = int(yv)

        if not mo:
            skip_no_month += 1
            if skip_no_month <= 3:
                logger.warning("B2B Daily: skipping row ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ cannot parse month from %r (type=%s) partner=%r",
                               raw_m, type(raw_m).__name__, partner)
            continue

        rev = _num(row_vals[rev_idx]) if rev_idx is not None and rev_idx < len(row_vals) else None
        pax = _num(row_vals[pax_idx]) if pax_idx is not None and pax_idx < len(row_vals) else None

        records.append({
            "sheet": "B2B Daily", "year": yr, "month": mo, "agency": partner,
            "revenue": round(rev * K_EUR, 2) if rev is not None else None,
            "pax": int(pax) if pax is not None else None,
            "plan": None, "ly": None,
        })

    logger.info("B2B Daily: %d records (skipped: no_partner=%d no_month=%d)", len(records), skip_no_partner, skip_no_month)
    return records


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ B2B Plan ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def parse_b2b_plan(raw_df: pd.DataFrame) -> List[Dict]:
    df = _promote_header(raw_df.copy())
    if df.empty:
        return []

    current_year = date.today().year

    month_map: Dict[str, int] = {}
    for col in df.columns:
        m = _col_to_month(col)
        if m:
            month_map[col] = m

    records = []

    if len(month_map) >= 2:
        def _fc(*kw):
            for col in df.columns:
                if any(k in str(col).lower() for k in kw):
                    return col
            return None
        partner_col = _fc("partner", "client", "agentie", "denumire", "name", "partener")
        for _, row in df.iterrows():
            partner = str(row[partner_col]).strip() if partner_col and pd.notna(row.get(partner_col)) else "TOTAL"
            if partner.lower() in ("nan", "none", ""):
                partner = "TOTAL"
            for col, m in month_map.items():
                v = _num(row.get(col))
                if v is None:
                    continue
                records.append({
                    "sheet": "B2B P", "year": current_year, "month": m,
                    "agency": partner, "plan": round(v * K_EUR, 2),
                    "revenue": None, "ly": None, "pax": None,
                })
    else:
        def _fc(*kw):
            for col in df.columns:
                if any(k in str(col).lower() for k in kw):
                    return col
            return None
        month_col   = _fc("luna", "month", "lunÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ", "period", "mon")
        plan_col    = _fc("plan", "target", "buget", "budget", "forecast", "value", "valoare")
        partner_col = _fc("partner", "client", "agentie", "denumire", "name")
        for _, row in df.iterrows():
            mo = _col_to_month(row.get(month_col)) if month_col else None
            if not mo:
                continue
            v = _num(row.get(plan_col)) if plan_col else None
            if v is None:
                continue
            partner = str(row[partner_col]).strip() if partner_col and pd.notna(row.get(partner_col)) else "TOTAL"
            records.append({
                "sheet": "B2B P", "year": current_year, "month": mo,
                "agency": partner, "plan": round(v * K_EUR, 2),
                "revenue": None, "ly": None, "pax": None,
            })

    logger.info("B2B Plan: %d records", len(records))
    return records




# ÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂ B2B LY ÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂ

def parse_b2b_ly(raw_df: pd.DataFrame, year: int) -> List[Dict]:
    """Parse B2B LY sheet ÃÂÃÂ¢ÃÂÃÂÃÂÃÂ same wide format as B2B P but stored as revenue for given year."""
    df = _promote_header(raw_df.copy())
    if df.empty:
        return []

    month_map: Dict[str, int] = {}
    for col in df.columns:
        m = _col_to_month(col)
        if m:
            month_map[col] = m

    records = []

    if len(month_map) >= 2:
        def _fc(*kw):
            for col in df.columns:
                if any(k in str(col).lower() for k in kw):
                    return col
            return None
        partner_col = _fc("partner", "client", "agentie", "denumire", "name", "partener")
        for _, row in df.iterrows():
            partner = str(row[partner_col]).strip() if partner_col and pd.notna(row.get(partner_col)) else "TOTAL"
            if partner.lower() in ("nan", "none", ""):
                partner = "TOTAL"
            for col, m in month_map.items():
                v = _num(row.get(col))
                if v is None:
                    continue
                records.append({
                    "sheet": "B2B LY", "year": year, "month": m,
                    "agency": partner, "revenue": round(v * K_EUR, 2),
                    "plan": None, "ly": None, "pax": None,
                })
    else:
        def _fc(*kw):
            for col in df.columns:
                if any(k in str(col).lower() for k in kw):
                    return col
            return None
        month_col   = _fc("luna", "month", "lun\u0103", "period", "mon")
        val_col     = _fc("ly", "actual", "realizat", "value", "valoare", "revenue", "vanzari")
        partner_col = _fc("partner", "client", "agentie", "denumire", "name")
        for _, row in df.iterrows():
            mo = _col_to_month(row.get(month_col)) if month_col else None
            if not mo:
                continue
            v = _num(row.get(val_col)) if val_col else None
            if v is None:
                continue
            partner = str(row[partner_col]).strip() if partner_col and pd.notna(row.get(partner_col)) else "TOTAL"
            records.append({
                "sheet": "B2B LY", "year": year, "month": mo,
                "agency": partner, "revenue": round(v * K_EUR, 2),
                "plan": None, "ly": None, "pax": None,
            })

    logger.info("B2B LY: %d records (year=%d)", len(records), year)
    return records

# ÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂ merge helpers ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def _merge_b2c(actuals, plan, ly, daily) -> List[Dict]:
    key_map: Dict[Tuple, Dict] = {}

    def _upsert(rec, fields):
        k = (rec["year"], rec["month"], rec.get("branch", ""))
        if k not in key_map:
            key_map[k] = {
                "sheet": rec.get("sheet", ""),
                "year": rec["year"], "month": rec["month"],
                "branch": rec.get("branch"), "region": rec.get("region"),
                "revenue": None, "plan": None, "ly": None,
                "pax": None, "reservations": None,
            }
        for f in fields:
            if rec.get(f) is not None:
                key_map[k][f] = rec[f]

    for r in actuals:
        _upsert(r, ["revenue", "region"])
    for r in plan:
        _upsert(r, ["plan", "region"])
    for r in ly:
        _upsert(r, ["ly", "region"])
    for r in daily:
        _upsert(r, ["pax", "reservations"])

    return list(key_map.values())


def _merge_b2b(daily, plan) -> List[Dict]:
    key_map: Dict[Tuple, Dict] = {}

    def _upsert(rec, fields):
        k = (rec["year"], rec["month"], rec.get("agency", ""))
        if k not in key_map:
            key_map[k] = {
                "sheet": rec.get("sheet", ""),
                "year": rec["year"], "month": rec["month"],
                "agency": rec.get("agency"),
                "revenue": None, "plan": None, "ly": None, "pax": None,
            }
        for f in fields:
            if rec.get(f) is not None:
                key_map[k][f] = rec[f]

    for r in daily:
        _upsert(r, ["revenue", "pax"])
    for r in plan:
        _upsert(r, ["plan"])

    return list(key_map.values())


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ master builder ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def build_wide_sheets(raw_bytes: bytes) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    PHASE 1 (fast, ~1-2 min): parse only the 4 small wide sheets.
    Returns (actuals, plan, ly, b2b_plan, valid_branches) without touching daily sheets.
    """
    import openpyxl

    current_year = date.today().year
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    available = wb.sheetnames
    logger.info("CT Dashboard.xlsx sheets: %s", available)

    def _load_df(sname: str) -> pd.DataFrame:
        if sname not in available:
            logger.warning("Sheet '%s' not in workbook", sname)
            return pd.DataFrame()
        rows = [list(r) for r in wb[sname].iter_rows(values_only=True)]
        df = pd.DataFrame(rows)
        del rows; gc.collect()
        logger.info("Loaded %s: %d rows ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ %d cols",
                    sname, len(df), len(df.columns) if not df.empty else 0)
        return df

    actuals_df = _load_df("B2C 2026 Actuals")
    actuals = parse_b2c_wide_sheet(actuals_df, "B2C 2026 Actuals", "revenue", current_year)
    del actuals_df; gc.collect()

    plan_df = _load_df("P")
    plan = parse_b2c_wide_sheet(plan_df, "P", "plan", current_year)
    del plan_df; gc.collect()

    ly_df = _load_df("LY")
    ly = parse_b2c_wide_sheet(ly_df, "LY", "ly", current_year)
    del ly_df; gc.collect()

    b2b_plan_df = _load_df("B2B P")
    b2b_plan = parse_b2b_plan(b2b_plan_df)
    del b2b_plan_df; gc.collect()

    b2b_ly_df = _load_df("B2B LY")
    b2b_ly = parse_b2b_ly(b2b_ly_df, current_year - 1)
    del b2b_ly_df; gc.collect()

    wb.close(); del wb; gc.collect()

    valid_branches: Set[str] = {r["branch"] for r in actuals + plan + ly if r.get("branch")}
    logger.info("Valid B2C branches: %d", len(valid_branches))
    return actuals, plan, ly, b2b_plan, b2b_ly, valid_branches



def _col_letter_to_idx(ref: str) -> int:
    """Convert cell ref like 'A1' or 'AB3' to 0-based column index."""
    idx = 0
    for ch in ref:
        if not ch.isalpha():
            break
        idx = idx * 26 + (ord(ch.upper()) - ord('A') + 1)
    return idx - 1


def _stream_b2b_daily_direct(xlsx_path: str) -> List[Dict]:
    """
    Stream B2B Daily directly from the xlsx ZIP using iterparse.
    Avoids loading the full openpyxl workbook ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ saves ~100MB+ RAM by
    streaming the shared strings and sheet XML element-by-element with
    elem.clear() after each row so memory stays flat.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    NS   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    try:
        zf = zipfile.ZipFile(xlsx_path, "r")
    except Exception as exc:
        logger.error("B2B Daily direct: cannot open zip %s: %s", xlsx_path, exc)
        return []

    try:
        names = zf.namelist()

        # ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ shared strings ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ iterparse to avoid loading whole XML ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in names:
            with zf.open("xl/sharedStrings.xml") as f:
                cur_parts: List[str] = []
                in_si = False
                for event, elem in ET.iterparse(f, events=["start", "end"]):
                    tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                    if event == "start" and tag == "si":
                        cur_parts = []
                        in_si = True
                    elif event == "end" and tag == "t" and in_si:
                        cur_parts.append(elem.text or "")
                    elif event == "end" and tag == "si":
                        shared_strings.append("".join(cur_parts))
                        in_si = False
                        elem.clear()
            logger.info("B2B Daily direct: %d shared strings loaded", len(shared_strings))

        # ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ find B2B Daily sheet path via workbook rels ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ
        sheet_rid = None
        with zf.open("xl/workbook.xml") as f:
            for event, elem in ET.iterparse(f, events=["end"]):
                tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                if tag == "sheet" and elem.get("name") == "B2B Daily":
                    for attr_key, attr_val in elem.attrib.items():
                        if attr_key.endswith("}id"):
                            sheet_rid = attr_val
                            break
                    if sheet_rid is None:
                        sheet_rid = elem.get(f"{{{NS_R}}}id")
                elem.clear()

        if sheet_rid is None:
            logger.warning("B2B Daily sheet not found in workbook.xml")
            return []

        sheet_path = None
        rels_path = "xl/_rels/workbook.xml.rels"
        if rels_path in names:
            with zf.open(rels_path) as f:
                for event, elem in ET.iterparse(f, events=["end"]):
                    if elem.get("Id") == sheet_rid:
                        target = elem.get("Target", "")
                        sheet_path = ("xl/" + target) if not target.startswith("xl/") else target
                    elem.clear()

        if not sheet_path or sheet_path not in names:
            logger.warning("B2B Daily: rels lookup failed (rid=%s path=%s) ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ trying index fallback",
                           sheet_rid, sheet_path)
            ws_files = sorted(n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"))
            sheet_path = ws_files[4] if len(ws_files) >= 5 else (ws_files[-1] if ws_files else None)
            if not sheet_path or sheet_path not in names:
                logger.error("B2B Daily direct: cannot locate worksheet XML")
                return []

        logger.info("B2B Daily direct: parsing %s", sheet_path)

        # ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ stream rows with iterparse, elem.clear() after each row ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ
        header = None
        partner_idx = pax_idx = rev_idx = month_idx = year_idx = None
        records: List[Dict] = []
        skip_no_partner = skip_no_month = 0
        current_year = date.today().year

        with zf.open(sheet_path) as f:
            for event, elem in ET.iterparse(f, events=["end"]):
                tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                if tag != "row":
                    continue

                # Build sparse row: column letter ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ value
                cells: dict = {}
                for c_elem in elem.findall(f"{{{NS}}}c"):
                    ref = c_elem.get("r", "")
                    if not ref:
                        continue
                    col_idx = _col_letter_to_idx(ref)
                    t_type  = c_elem.get("t", "")
                    v_elem  = c_elem.find(f"{{{NS}}}v")
                    if v_elem is None:
                        val = None
                    elif t_type == "s":
                        try:
                            val = shared_strings[int(v_elem.text)]
                        except (IndexError, ValueError, TypeError):
                            val = None
                    elif t_type in ("str", "inlineStr"):
                        val = v_elem.text
                    else:
                        try:
                            val = float(v_elem.text)
                        except (ValueError, TypeError):
                            val = v_elem.text
                    cells[col_idx] = val

                elem.clear()  # free XML memory immediately

                if not cells:
                    continue

                max_col  = max(cells.keys()) + 1
                row_vals = [cells.get(i) for i in range(max_col)]

                if not any(v is not None for v in row_vals):
                    continue

                if header is None:
                    str_count = sum(1 for v in row_vals if isinstance(v, str) and str(v).strip())
                    if str_count >= 2:
                        candidate = [str(v).strip().lower() if v is not None else "" for v in row_vals]
                        c_partner = _header_col_idx(candidate, "partner", "client", "agentie",
                                                     "agency", "denumire", "name", "partener")
                        c_month   = _header_col_idx(candidate, "luna", "month", "lunÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ",
                                                     "period", "data", "date")
                        if c_partner is not None and c_month is not None:
                            header      = candidate
                            partner_idx = c_partner
                            pax_idx     = _header_col_idx(header, "pax", "pasageri", "persons")
                            rev_idx     = _header_col_idx(header, "valoare", "value", "revenue",
                                                          "vanzari", "actual", "sales",
                                                          "incastat", "eur", "nett", "net",
                                                          "total", "realizat",
                                                          exclude={pax_idx} if pax_idx is not None else None)
                            month_idx   = c_month
                            year_idx    = _header_col_idx(header, "year", "an", "anul")
                            logger.info(
                                "B2B Daily direct header: partner=%s pax=%s rev=%s month=%s | cols=%s",
                                partner_idx, pax_idx, rev_idx, month_idx,
                                [str(h)[:20] if h else None for h in row_vals[:8]])
                            _b2b_stream_diag["detected"] = {
                                "partner_idx": partner_idx, "pax_idx": pax_idx,
                                "rev_idx": rev_idx, "month_idx": month_idx,
                                "header_row": [str(h)[:30] if h else None for h in row_vals[:10]]
                            }
                            _b2b_stream_diag["sample_rows"] = []
                        else:
                            logger.info("B2B Daily direct: skipping non-header row: %s",
                                        [str(v)[:15] for v in row_vals[:6]])
                    continue

                # parse partner
                partner = None
                if partner_idx is not None and partner_idx < len(row_vals) and row_vals[partner_idx] is not None:
                    partner = str(row_vals[partner_idx]).strip()
                if len(_b2b_stream_diag.get("sample_rows", [])) < 5:
                    _b2b_stream_diag.setdefault("sample_rows", []).append(
                        [str(v)[:25] if v is not None else None for v in row_vals[:8]])
                if not partner or partner.lower() in ("nan", "none", "", "-"):
                    skip_no_partner += 1
                    continue
                if partner.replace(".", "").lstrip("-").isdigit():
                    skip_no_partner += 1
                    continue

                # parse month/year
                raw_m = row_vals[month_idx] if month_idx is not None and month_idx < len(row_vals) else None
                mo = yr = None

                if isinstance(raw_m, (datetime, date)):
                    mo, yr = raw_m.month, raw_m.year
                elif isinstance(raw_m, (int, float)) and 45 < raw_m < 100000:
                    # Excel serial date — extract both month AND year
                    try:
                        _dt = date(1899, 12, 30) + timedelta(days=int(raw_m))
                        mo, yr = _dt.month, _dt.year
                    except Exception:
                        pass
                elif raw_m is not None:
                    mo = _col_to_month(raw_m)
                    if not mo:
                        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%m/%Y", "%d/%m/%Y"):
                            try:
                                dt_parsed = datetime.strptime(str(raw_m).strip(), fmt)
                                mo, yr = dt_parsed.month, dt_parsed.year
                                break
                            except ValueError:
                                pass

                # Try to extract year from month string (e.g. "Ian 2025", "January 2025")
                if yr is None and raw_m is not None:
                    import re as _re
                    _m_yr = _re.search(r'\b(20\d\d)\b', str(raw_m))
                    if _m_yr:
                        yr = int(_m_yr.group(1))
                if yr is None:
                    yr = current_year
                    if year_idx is not None and year_idx < len(row_vals):
                        yv = _num(row_vals[year_idx])
                        if yv:
                            yr = int(yv)

                if not mo:
                    skip_no_month += 1
                    if skip_no_month <= 3:
                        logger.warning(
                            "B2B Daily direct: skipping row ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ cannot parse month from %r (type=%s) partner=%r",
                            raw_m, type(raw_m).__name__, partner)
                    continue

                rev = _num(row_vals[rev_idx]) if rev_idx is not None and rev_idx < len(row_vals) else None
                pax = _num(row_vals[pax_idx]) if pax_idx is not None and pax_idx < len(row_vals) else None

                records.append({
                    "sheet": "B2B Daily", "year": yr, "month": mo, "agency": partner,
                    "revenue": round(rev * K_EUR, 2) if rev is not None else None,
                    "pax": int(pax) if pax is not None else None,
                    "plan": None, "ly": None,
                })

        logger.info("B2B Daily direct: %d records (skipped: no_partner=%d no_month=%d)",
                    len(records), skip_no_partner, skip_no_month)
        return records

    except Exception as exc:
        logger.error("B2B Daily direct failed: %s", exc, exc_info=True)
        return []
    finally:
        try:
            zf.close()
        except Exception:
            pass


def merge_daily_into_cache(xlsx_path: str, valid_branches: Set[str],
                            b2c_ts: List[Dict], b2b_ts: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """
    PHASE 2 (slow): stream B2B Daily via direct ZIP streamer (no openpyxl, no OOM).
    B2C actuals are already loaded in Phase 1 from the 'B2C 2026 Actuals' wide sheet.
    Returns (b2c_ts unchanged, b2b_ts + b2b_daily_records).
    """
    logger.info("Phase 2 starting...")
    try:
        b2b_daily = _stream_b2b_daily_direct(xlsx_path)
        logger.info("Phase 2: streamed %d B2B daily records", len(b2b_daily))
        if not b2b_daily:
            logger.warning("Phase 2: no B2B daily records found - returning unchanged caches")
            return b2c_ts, b2b_ts
        b2b_updated = b2b_ts + b2b_daily
        logger.info("Phase 2 done - B2C: %d  B2B: %d records", len(b2c_ts), len(b2b_updated))
        return b2c_ts, b2b_updated
    except Exception as e:
        logger.error("Phase 2 failed: %s", e, exc_info=True)
        return b2c_ts, b2b_ts

def build_dashboard(raw_bytes: bytes) -> Tuple[List[Dict], List[Dict]]:
    """Convenience wrapper: phase 1 only (fast). Phase 2 handled separately in main.py."""
    actuals, plan, ly, b2b_plan, b2b_ly, _ = build_wide_sheets(raw_bytes)
    b2c = _merge_b2c(actuals, plan, ly, [])
    b2b = _merge_b2b(b2b_ly, b2b_plan)
    logger.info("Wide-sheet build ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ B2C: %d  B2B: %d records", len(b2c), len(b2b))
    return b2c, b2b


# ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ debug helper ÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ¢ÃÂÃÂÃÂÃÂÃÂÃÂÃÂÃÂ

def sheet_raw_rows(raw_bytes: bytes, sheet_name: str, nrows: int = 8) -> dict:
    """Return first nrows of a sheet as plain dicts for the debug endpoint."""
    import openpyxl
    wb = openpyxl.load_workbook(
        io.BytesIO(raw_bytes), read_only=True, data_only=True
    )
    try:
        if sheet_name not in wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found",
                    "available": wb.sheetnames}
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= nrows:
                break
            rows.append([str(v) if v is not None else "" for v in row])
        return {"sheet": sheet_name, "rows": rows}
    finally:
        wb.close()


def sheet_names(raw_bytes: bytes) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names
