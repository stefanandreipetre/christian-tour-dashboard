"""
Christian Tour Sales Dashboard — FastAPI Backend
Single source: CT Dashboard.xlsx (SharePoint)
Two-phase load:
  Phase 1 (fast): wide sheets -> cache immediately (~2 min)
  Phase 2 (slow): stream daily sheets -> merge into cache (background)
"""

import os
import gc
import logging
import time
import threading
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv

import cache
import data_processor as dp
from sharepoint_client import SharePointClient

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s - %(message)s")
logger = logging.getLogger(__name__)

sp        = SharePointClient()
scheduler = BackgroundScheduler()
REFRESH_INTERVAL = int(os.getenv("REFRESH_INTERVAL", 3600))
_load_lock = threading.Lock()


def load_dashboard() -> None:
    """
    Two-phase build so the cache is populated quickly:
      Phase 1: download + parse wide sheets -> set cache (visible within ~2 min)
      Phase 2: stream daily sheets -> merge pax/res into existing cache
    """
    with _load_lock:
        try:
            # Phase 1: small/wide sheets
            logger.info("Downloading CT Dashboard.xlsx ...")
            raw = sp.download_file()
            logger.info("Downloaded %d bytes - parsing wide sheets ...", len(raw))

            actuals, plan, ly, b2b_plan, valid_branches = dp.build_wide_sheets(raw)

            b2c_ts = dp._merge_b2c(actuals, plan, ly, [])
            b2b_ts = dp._merge_b2b([], b2b_plan)

            cache.set_data("b2c", None, b2c_ts)
            cache.set_data("b2b", None, b2b_ts)
            logger.info("Phase 1 done - B2C: %d  B2B: %d records (wide sheets only)",
                        len(b2c_ts), len(b2b_ts))

            # Phase 2: stream daily sheets
            logger.info("Phase 2: streaming daily sheets ...")
            try:
                b2c_updated, b2b_updated = dp.merge_daily_into_cache(
                    raw, valid_branches, b2c_ts, b2b_ts
                )
                del raw; gc.collect()

                cache.set_data("b2c", None, b2c_updated)
                cache.set_data("b2b", None, b2b_updated)
                logger.info("Phase 2 done - B2C: %d  B2B: %d records (with daily)",
                            len(b2c_updated), len(b2b_updated))
            except Exception as exc:
                logger.error("Phase 2 (daily streaming) failed - keeping wide-sheet cache: %s",
                             exc, exc_info=True)
                try:
                    del raw
                except NameError:
                    pass
                gc.collect()

        except Exception as exc:
            logger.error("load_dashboard failed: %s", exc, exc_info=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    threading.Thread(target=load_dashboard, daemon=True).start()
    scheduler.add_job(load_dashboard, "interval", seconds=REFRESH_INTERVAL,
                      id="refresh", replace_existing=True)
    scheduler.start()
    yield
    scheduler.shutdown(wait=False)


app = FastAPI(title="CT Dashboard API", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


def _ts(channel: str):
    data = cache.get_data(channel)
    if data is None:
        raise HTTPException(503, f"{channel} not yet loaded - retry in a moment")
    return data.get("timeseries", [])


def _filter(records, year=None, month=None):
    out = records
    if year:
        out = [r for r in out if r.get("year") == year]
    if month:
        out = [r for r in out if r.get("month") == month]
    return out


@app.get("/api/health")
def health():
    return {"status": "ok", "time": time.time()}


@app.get("/api/refresh")
def manual_refresh(background_tasks: BackgroundTasks):
    background_tasks.add_task(load_dashboard)
    return {"status": "refresh started"}


@app.get("/api/status")
def status():
    b2c = cache.get_data("b2c")
    b2b = cache.get_data("b2b")
    return {
        "b2c": {"records": len(b2c["timeseries"]) if b2c else 0,
                "updated_at": b2c["updated_at"] if b2c else None},
        "b2b": {"records": len(b2b["timeseries"]) if b2b else 0,
                "updated_at": b2b["updated_at"] if b2b else None},
    }


# B2C endpoints

@app.get("/api/b2c/monthly")
def b2c_monthly(year: Optional[int] = None, month: Optional[int] = None):
    return _filter(_ts("b2c"), year, month)


@app.get("/api/b2c/summary")
def b2c_summary(year: Optional[int] = None, month: Optional[int] = None):
    agg: dict = {}
    for r in _filter(_ts("b2c"), year, month):
        k = (r.get("year"), r.get("month"))
        if k not in agg:
            agg[k] = {"year": k[0], "month": k[1],
                      "revenue": 0.0, "plan": 0.0, "ly": 0.0, "pax": 0, "reservations": 0}
        for f in ("revenue", "plan", "ly"):
            if r.get(f) is not None:
                agg[k][f] = round(agg[k][f] + r[f], 2)
        for f in ("pax", "reservations"):
            if r.get(f) is not None:
                agg[k][f] += r[f]
    return sorted(agg.values(), key=lambda x: (x["year"], x["month"]))


@app.get("/api/b2c/branches")
def b2c_branches(year: Optional[int] = None):
    agg: dict = {}
    for r in _filter(_ts("b2c"), year):
        b = r.get("branch")
        if not b:
            continue
        if b not in agg:
            agg[b] = {"branch": b, "region": r.get("region", ""), "revenue": 0.0, "plan": 0.0, "ly": 0.0, "pax": 0}
        for f in ("revenue", "plan", "ly"):
            if r.get(f) is not None:
                agg[b][f] = round(agg[b][f] + r[f], 2)
        if r.get("pax") is not None:
            agg[b]["pax"] += int(r["pax"])
    result = sorted(agg.values(), key=lambda x: x["revenue"], reverse=True)
    for r in result:
        r["bookings"] = r["pax"]
        r["vs_plan_pct"] = round((r["revenue"] / r["plan"] - 1) * 100, 1) if r.get("plan") else None
    return result


# B2B endpoints

@app.get("/api/b2b/monthly")
def b2b_monthly(year: Optional[int] = None, month: Optional[int] = None):
    return _filter(_ts("b2b"), year, month)


@app.get("/api/b2b/summary")
def b2b_summary(year: Optional[int] = None, month: Optional[int] = None):
    agg: dict = {}
    for r in _filter(_ts("b2b"), year, month):
        k = (r.get("year"), r.get("month"))
        if k not in agg:
            agg[k] = {"year": k[0], "month": k[1],
                      "revenue": 0.0, "plan": 0.0, "ly": 0.0, "pax": 0}
        for f in ("revenue", "plan", "ly"):
            if r.get(f) is not None:
                agg[k][f] = round(agg[k][f] + r[f], 2)
        if r.get("pax") is not None:
            agg[k]["pax"] += r["pax"]
    return sorted(agg.values(), key=lambda x: (x["year"], x["month"]))


@app.get("/api/b2b/partners")
def b2b_partners(year: Optional[int] = None):
    agg: dict = {}
    for r in _filter(_ts("b2b"), year):
        a = r.get("agency")
        if not a:
            continue
        if a not in agg:
            agg[a] = {"agency": a, "pax": 0}
        pax_val = int(r.get("pax") or r.get("revenue") or 0)
        agg[a]["pax"] += pax_val
    result = sorted(agg.values(), key=lambda x: x["pax"], reverse=True)
    for r in result:
        r["bookings"] = r["pax"]
        r["revenue"] = r["pax"]  # PAX as proxy revenue for display compatibility
    return result


# Debug endpoints

@app.get("/api/debug/sheets")
def debug_sheets():
    try:
        raw = sp.download_file()
        names = dp.sheet_names(raw)
        return {"sheets": names}
    except Exception as exc:
        return {"error": str(exc)}


@app.get("/api/debug/sheet-raw")
def debug_sheet_raw(sheet: str, nrows: int = 8):
    try:
        raw = sp.download_file()
        return dp.sheet_raw_rows(raw, sheet, nrows)
    except Exception as exc:
        return {"error": str(exc)}


@app.get("/api/debug/b2b-daily-shape")
def debug_b2b_daily():
    """Return raw headers + first 5 rows of B2B Daily sheet for diagnosis."""
    import openpyxl, io
    raw = cache._raw
    if raw is None:
        return {"error": "cache not loaded yet"}
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if "B2B Daily" not in wb.sheetnames:
        return {"error": "B2B Daily sheet not found", "sheets": wb.sheetnames}
    ws = wb["B2B Daily"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 8:
            break
        rows.append([str(v) if v is not None else None for v in row])
    return {"rows": rows, "total_rows_sampled": len(rows)}
